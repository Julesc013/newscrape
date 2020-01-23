import os
import subprocess
from shutil import copyfile
from openpyxl import Workbook
from openpyxl import load_workbook,styles
import requests
from lxml import html
from bs4 import BeautifulSoup
import unicodecsv as csv
import argparse
from requests_testadapter import Resp

class LocalFileAdapter(requests.adapters.HTTPAdapter): # This allows a locally saved HTML file to be retrieved just as the "get" function would.
    def build_response_from_file(self, request):
        file_path = request.url[7:]
        with open(file_path, 'rb') as file:
            buff = bytearray(os.path.getsize(file_path))
            file.readinto(buff)
            resp = Resp(buff)
            r = self.build_response(request, resp)

            return r

    def send(self, request, stream=False, timeout=None,
             verify=True, cert=None, proxies=None):

        return self.build_response_from_file(request)

class listing: # Record structure to hold new listings found on each page.
    def __init__(self, business_name, phone_number, email_address, business_website):
        self.business_name = business_name
        self.phone_number = phone_number
        self.email_address = email_address
        self.business_website = business_website

def search_data(sheet, column, text): # Search the existing data for matches... if found, don't count this as a new listing # Sheet must be an Excel worksheet, Column should be in set {A,B,C,D} and Text should be alphanumeric.

    # Get all cells from specified column
    for cell in sheet[column]:
        if(cell.value is not None): # We need to check that the cell is not empty
            if text in cell.value: # Check if the value of the cell contains the text of the business phone number
                return True # return true then break out of this function
    
    return False # Else (if none found) return false



# Define variables

pages_path = r"/home/webscraper/Documents/Newscrape/Pages/"
results_path = r"/home/webscraper/Documents/Newscrape/Results/"
all_path = results_path + "all_results.xlsx"
new_path = results_path + "new_results.xlsx"
template_path = results_path + "new_results_template.xlsx"


# Make a fresh copy of the new-results template for editing.

print("Copying spreadsheet template...", end="")

copyfile(template_path, new_path)

print(" Done.")


# Load worksheets

print("Loading worksheets...", end="")

book_all = load_workbook(filename = all_path) # Load the workbook
sheet_all = book_all['Sheet1'] # Load the worksheet

book_new = load_workbook(filename = new_path) # Load the workbook
sheet_new = book_new['Sheet1'] # Load the worksheet

print(" Done.")


####################################https://www.pluralsight.com/guides/extracting-data-html-beautifulsoup

pages_bytes = os.fsencode(pages_path) # Get the directory link
for page_bytes in os.listdir(pages_bytes):

    # Declare list that will hold listing data.
    listings = []

    # Get file name and path
    page_file = os.fsdecode(page_bytes)
    page_path = pages_path + page_file

    print("Getting " + page_file + "...", end="") # DEBUG
    
    requests_session = requests.session()
    requests_session.mount('file://', LocalFileAdapter())
    response = requests_session.get('file://' + page_path)

    # Check that the page parsed correctly before processing the data.
    if response.status_code==200:

        print(" Success.")
        print("Parsing HTML code... ", end="")

        soup = BeautifulSoup(response.text, 'html.parser')

        print(" Done.")
        print("Finding listings data... ", end="")


        # Gather all listings' data and check if they have already been identified.

        # Get the html data of each match.
        business_names_html = soup.find_all("a", attrs={"class": "listing-name"})
        phone_numbers_html = soup.find_all("a", attrs={"class": "click-to-call contact contact-preferred contact-phone"})
        email_addresses_html = soup.find_all("a", attrs={"class": "contact contact-main contact-email"})
        business_websites_html = soup.find_all("a", attrs={"class": "class='contact contact-main contact-url"})

        # DEBUG PRINT LENGTHS
        print(len(business_names_html))
        print(len(phone_numbers_html))
        print(len(email_addresses_html))
        print(len(business_websites_html))


        # Extract the appropriate sections of each html element.
        number_of_listings = len(business_names_html)
        for index in range(0, number_of_listings - 1):

            print(index) #DEBUG

            # Get sections
            name = business_names_html[index].get('text')
            phone = phone_numbers_html[index].get('href')
            email = email_addresses_html[index].get('data-email')
            website = business_websites_html[index].get('href')

            # Add sections to a record
            record = listing(name, phone, email, website)
            # Append this record to the list of listings
            listings.append(record)

        print(" Done.")

        #TEST PRINT DEBUG
        print(listings)


        # Check if this business listing exists


    else:
        print(" Failed. Error code: " + response.status_code)
        