import os
from shutil import copyfile
from openpyxl import Workbook
from openpyxl import load_workbook,styles
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from math import ceil
from datetime import datetime
import list_data # The file containing the lists of suburbs (in the same directory)


class listing: # Record structure to hold new listings found on each page.
    def __init__(self, business_name, phone_number, email_address, business_website, yellow_pages_link):
        self.business_name = business_name
        self.phone_number = phone_number
        self.email_address = email_address
        self.business_website = business_website
        self.yellow_pages_link = yellow_pages_link


def get_time_now():

    time = datetime.now()
    time_string = "[" + time.strftime("%d/%m/%Y %H:%M:%S") + "]"
    return time_string

def write_to_log(output):
    with open(log_file_path, 'a') as log_file:
        log_file.write(output)

def console_action(action, details):
    # E.g. Saving file...
    # No newline, first word yellow, has timestamp.

    if details != "":
        output = get_time_now() + " " + action + " " + details + "..."
    else:
        output = get_time_now() + " " + action + "..."

    print(output, end="")
    write_to_log(output)

def console_complete(result, status): # Status is a boolean representing success
    # E.g. Done.
    # Newline, green or red.

    output = result + "."
    print(" " + output)
    write_to_log(output + "\n")

def console_message(message):
    
    # Newline, magenta.
    output = get_time_now() + " " + message + "."
    print(output)
    write_to_log(output + "\n")


def find_match(sheet, column, text): # Search the existing data for matches... if found, don't count this as a new listing # Sheet must be an Excel worksheet, Column should be in set {A,B,C,D,E} and Text should be alphanumeric. # Returns True if a match is found.

    if text is not None: # Check that the string provided exists

        # Get all cells from specified column
        for cell in sheet[column]:

            if(cell.value is not None): # We need to check that the cell is not empty

                if cell.value == text: # Check if the value of the cell contains the text

                    return True # return true then break out of this function
        
    return False # Else (if none found) return false



# Define variables

version = "1.1.0"

browser = webdriver.Firefox()
address_base = ("https://www.yellowpages.com.au/search/listings?clue=", "&eventType=pagination&openNow=false&pageNumber=", "&referredBy=UNKNOWN&&state=", "&suburb=", "+") # [0], clue, [1], page, [2] state, [3] suburb+state)

clues = ('electricians+electrical+contractors', 'plumbers+gas+fitters', 'builders+building+contractors')
states = ('NSW', 'VIC', 'QLD', 'ACT', 'SA', 'WA', 'NT', 'TAS')

logs_path = r"/home/webscraper/Documents/Newscrape/Logs/"
results_path = r"/home/webscraper/Documents/Newscrape/Results/"

all_path = results_path + "all_results.xlsx"
new_path = results_path + "new_results.xlsx"
template_path = results_path + "new_results_template.xlsx"

# Declare lists that will hold listing records.
listings = []
new_listings = []

# Get the list of suburbs (as a disctionary of tuples)
suburbs = list_data.get_suburbs()


# BEGIN ACTIONS


# Make a new log file for this session

if not os.path.exists(logs_path): # Create the directories if they don't exist
    os.makedirs(logs_path)

time_atm = datetime.now()
time_atm_string = time_atm.strftime("%d%m%Y-%H%M%S")
log_file_name = "newscrape-" + time_atm_string + ".log"
log_file_path = os.path.join(logs_path, log_file_name)

with open(log_file_path, 'w') as log_file:
    log_file.write('Newscrape Console Log ' + get_time_now() + " (Version " + version + ")\n") # Write the header

console_message("Created new log file " + log_file_name)



# Count total pages to search through

total_clues = len(clues)
pages_multiplier = 1.0185 # Add 1.85% (derived from test data)
time_per_search = 10.3 # On average 10.3002 seconds per search (time per clue is 16.75hrs)
time_per_check = 0.1 # NOT A REAL VALUE, ONLY AN ESTIMATE, REPLACE LATER
#time_per_rank = 10 #TEMPVAR (ASIC RANKING)
checks_per_clue = 25000 # A really rough average
#ranks_per_clue = 100 #TEMPVAR (ASIC RANKING)

total_suburbs = 0
for state in states:
    total_suburbs += len(suburbs[state]) # Add up all the suburbs
total_searches = total_suburbs * total_clues # Multiply by the number of times we have to fetch from each suburb
total_checks = checks_per_clue * total_clues

expected_duration = total_searches * time_per_search + total_checks * time_per_check # Calculate the expected total duration of the program
expected_duration_timedelta = datetime.timedelta(seconds=expected_duration)

start_time = datetime.now

# Print this information and time estimates

print("Total suburbs to search: " + str(total_suburbs), \
    "Total searches to do: " + str(total_searches), \
    "Expected total pages to search: " + str(total_searches * pages_multiplier), \
    "Expected run duration: " + str(expected_duration_timedelta), \
    "Expected time of completion: " + str(start_time + expected_duration_timedelta), \
    sep="\n")



# Make a fresh copy of the new-results template for editing.
console_action("Copying spreadsheet template", "")
copyfile(template_path, new_path)
console_complete("Done", True)



# Download and extract data from every page!

# Initialise counting variables
suburbs_counter = 0
pages_counter = 0

# .split()[0]Loop through each, clue, state, suburb, and page.

for clue in clues:

    for state in states:

        for suburb in suburbs[state]: # Only the suburbs in this state

            suburbs_counter += 1

            # Reset page numbers file
            pages_number = -1

            # Loop through all pages
            page = 1
            while ( (page <= pages_number or pages_number == -1) and page <= 8):

                pages_counter += 1

                # Construct the url
                web_address = address_base[0] + clue + address_base[1] + str(page) + address_base[2] + state + address_base[3] + suburb + address_base[4] + state

                # Get the html for this page
                while True: # Loop forever until successful
                    try:

                        percentage_complete = (suburbs_counter / total_suburbs) * 100

                        console_action("[" + str(round(percentage_complete, 2)) + "%] " + "Getting", web_address)
                
                        browser.get(web_address)
                        html_source = browser.page_source

                        console_complete("Done", True)
                        break

                    except:

                        console_complete("Failed", False)
                        pass

                console_action("Parsing and extracting HTML code", "")

                soup = BeautifulSoup(html_source, 'html.parser')
                listings_html = soup.find_all("div", attrs={"class": "listing listing-search listing-data"})

                console_complete("Done", True)

                # If the number of pages hasn't been retrieved, get it from the page just downloaded
                if pages_number == -1:

                    console_action("Finding number of pages", "")

                    try:
                            
                        # Extract the number of results
                        results_number_parent = soup.find("div", attrs={"class": "cell search-message first-cell"})
                        results_number_html = results_number_parent.find("span", attrs={"class": "emphasise"})
                        results_number = int(results_number_html.text.split()[0])

                        # Do the math
                        if results_number >= 1 and results_number <= 1500:

                            pages_number = int(ceil(results_number / 35.0))

                            console_complete("Done", True)
                            
                        else:

                            pages_number = -1

                            console_complete("Unreasonable result", False)
                        
                    except:

                        pages_number = 0

                        console_complete("Failed (assuming zero results)", False)


                # Gather all listings' data and check if they have already been identified.

                for listing_html in listings_html:

                    console_action("Extracting listing", listing_html.get('data-listing-name'))

                    ## Make links absolute
                    base_url = "https://www.yellowpages.com.au"
                    #listing_html.make_links_absolute(base_url)


                    # Reset variables
                    business_name_html = None
                    phone_number_html = None
                    email_address_html = None
                    business_website_html = None
                    name = None
                    phone = None
                    email = None
                    website = None
                    yellow_page = None

                    # Get the html data of each match.
                    business_name_html = listing_html.find("a", attrs={"class": "listing-name"})
                    #phone_number_html = listing_html.find("a", attrs={"class": "click-to-call contact contact-preferred contact-phone"})
                    phone_number_html = listing_html.find("a", attrs={"class": lambda x: x and 'click-to-call contact contact-preferred' in x})
                    email_address_html = listing_html.find("a", attrs={"class": "contact contact-main contact-email"})
                    business_website_html = listing_html.find("a", attrs={"class": "contact contact-main contact-url"})


                    # Extract the appropriate sections of each html element.

                    # Get sections (only if they exist)
                    if business_name_html is not None:
                        name = business_name_html.text
                    if phone_number_html is not None:
                        phone = phone_number_html.get('href')
                    if email_address_html is not None:
                        email = email_address_html.get('data-email')
                    if business_website_html is not None:
                        website = business_website_html.get('href')
                    if business_name_html is not None:
                        yellow_page = business_name_html.get('href') # Link to the Yellow Pages listing.

                    # Add sections to a record
                    record = listing(name, phone, email, website, base_url + yellow_page)
                    #record = listing(name, phone, email, base_url + yellow_page) # Use the YP listing instead of the actual website temporarily.
                    # Append this record to the list of listings
                    listings.append(record)


                    console_complete("Done", True)


                page += 1 # Increment page number


time_searching_done = datetime.now # The time at which all the downloading and searching was completed


# Check if each business listing already exists in our local database


# Load all-results worksheet

console_action("Loading worksheet", "all-results")

book_all = load_workbook(filename = all_path) # Load the workbook
sheet_all = book_all['Sheet1'] # Load the worksheet

console_complete("Done", True)


# Get the last row in the sheet
final_row_all = sheet_all.max_row


# Loop through each newly retrieved record

index = 0
sheet_index = 0 # This is used so that no rows are skipped in the spreadsheet
listings_count = len(listings)
while index <= listings_count - 1:

    business = listings[index]

    console_action("Checking", business.business_name)

    this_name = business.business_name
    this_phone = business.phone_number
    this_email = business.email_address
    this_website = business.business_website
    this_yellow_page = business.yellow_pages_link

    # Search the sheet of all listings to see if it already exists
    this_name_exists = find_match(sheet_all, "A", this_name)
    this_phone_exists = find_match(sheet_all, "B", this_phone)
    this_email_exists = find_match(sheet_all, "C", this_email)
    this_website_exists = find_match(sheet_all, "D", this_website)
    this_yellow_page_exists = find_match(sheet_all, "E", this_yellow_page)

    if this_name_exists or this_phone_exists or this_email_exists or this_website_exists or this_yellow_page_exists: # If any of the searches returned a True result for existence

        console_complete("Already exists", False)

        # Keep sheet_index the same

    else:

        console_complete("Found new listing", True)

        new_listings.append(business) # Add this new listing to the list of new listings

        # Add the new listing to all-results spreadsheet

        # Using the old max row as a base get the next row number to write to
        this_row_all = final_row_all + sheet_index + 1

        console_action("Adding", business.business_name + " to spreadsheet all-results")

        # Add to all listings sheet
        sheet_all.cell(row=this_row_all, column=1).value = this_name # Name
        sheet_all.cell(row=this_row_all, column=2).value = this_phone # Phone
        sheet_all.cell(row=this_row_all, column=3).value = this_email # Email
        sheet_all.cell(row=this_row_all, column=4).value = this_website # Website
        sheet_all.cell(row=this_row_all, column=5).value = this_yellow_page # Yellow Page


        console_complete("Done", True)

        sheet_index += 1


    index += 1 # Increment index (b/c not using a for loop)


console_action("Saving spreadsheet", "results-all")

# Save the changes to the file
book_all.save(all_path)

console_complete("Done", True)


time_checking_done = datetime.now # The time at which all the checking was completed (incl. writing to the all-spreadsheet)


# Get data on new listings and add them to the new-reults spreadsheet

# Load new-results worksheet

console_action("Loading worksheet", "results-new")

book_new = load_workbook(filename = new_path) # Load the workbook
sheet_new = book_new['Sheet1'] # Load the worksheet

console_complete("Done", True)


# Get the last row in the sheet
final_row_new = 1 #sheet_new.max_row # Always start at the top of the sheet


# Loop through each newly retrieved record

index = 0
new_listings_count = len(new_listings)
while index <= new_listings_count - 1:

    new_business = new_listings[index]

    #console_action("Getting ASIC data for", new_business.business_name)

    ####### GET ASIC DATA!!!!!!!!!!!!!!!!!!

    this_name = new_business.business_name
    this_phone = new_business.phone_number
    this_email = new_business.email_address
    this_website = new_business.business_website
    this_yellow_page = new_business.yellow_pages_link

    # Add the new listing to the new-listings spreadsheet

    # Using the old max row as a base get the next row number to write to
    this_row_new = final_row_new + sheet_index + 1 # Start at the top

    console_action("Adding", new_business.business_name + " to spreadsheet results-new")

    # Add to new listings sheet
    sheet_new.cell(row=this_row_new, column=1).value = this_name # Name
    sheet_new.cell(row=this_row_new, column=2).value = this_phone # Phone
    sheet_new.cell(row=this_row_new, column=3).value = this_email # Email
    sheet_new.cell(row=this_row_new, column=4).value = this_website # Website
    sheet_new.cell(row=this_row_new, column=5).value = this_yellow_page # Yellow Page


    console_complete("Done", True)

    index += 1 # Increment index (b/c not using a for loop)


console_action("Saving spreadsheet", "results-new")

# Save the changes to the file
book_new.save(new_path)

console_complete("Done", True)


time_ranking_done = datetime.now # The time at which all the ranking (with ASIC) was completed (incl. writing to the new spreadsheet)



# GET ASIC (ABN/ACN) DETAILS and SORT NEW LISTINGS BASED ON IF THEY HAVE A WEBSITE/ABN/ACN ((SEE ABOVE--GOES INSIDE FOR LOOP)).
# ADD ASIC RANKING TO TIME CALCULATIONS

# EMAIL THE SHEETS.


# Calculate and print statistics

#total_pages_count = pages_counter
total_listings_count = len(listings)
new_listings_count = len(new_listings)
print("Total pages: " + str(pages_counter), \
    "Total listings: " + str(total_listings_count), \
    "Total new listings: " + str(new_listings_count), \
    sep="\n")

time_searching_duration = time_searching_done - start_time
time_checking_duration = time_checking_done - time_searching_done
time_ranking_duration = time_ranking_done - time_checking_done
print("Duration of searching: " + str(time_searching_duration), \
    "Duration of checking: " + str(time_checking_duration), \
    "Duration of ranking: " + str(time_ranking_duration), \
    sep="\n")

finish_time = datetime.now
total_duration = finish_time - start_time
difference_duration = total_duration - expected_duration
print("Total duration: " + str(total_duration), \
    "Expected duration: " + str(expected_duration), \
    "Difference from expected: " + str(difference_duration), \
    sep="\n")