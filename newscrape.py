import os
from subprocess import call
from shutil import copyfile
from openpyxl import Workbook
from openpyxl import load_workbook,styles
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
import openpyn
from math import ceil
import random
from time import time, sleep
from datetime import datetime, date, time, timedelta
import list_data # The file containing the lists of suburbs (in the same directory)
import smtplib 
from email.mime.multipart import MIMEMultipart 
from email.mime.text import MIMEText 
from email.mime.base import MIMEBase 
from email import encoders 


class listing: # Record structure to hold new listings found on each page.
    def __init__(self, business_name, phone_number, email_address, business_website, yellow_pages_link, location_state):
        self.business_name = business_name
        self.phone_number = phone_number
        self.email_address = email_address
        self.business_website = business_website
        self.yellow_pages_link = yellow_pages_link
        self.location_state = location_state


def get_time_now():

    time = datetime.now()
    time_string = "[" + time.strftime("%d/%m/%Y %H:%M:%S") + "]"
    return time_string

def format_time_seconds(time_seconds):

    # Take in the time in seconds and return the HMS format string
    minutes, seconds = divmod(time_seconds, 60)
    hours, minutes = divmod(minutes, 60)
    return f'{int(hours):d}:{int(minutes):02d}:{int(seconds):02d}'


def write_to_log(output):
    with open(log_file_path, 'a') as log_file:
        log_file.write(output)

def console_action(action, details):
    # E.g. Saving file...
    # No newline, first word yellow, has timestamp.

    if details != "":
        output = get_time_now() + " " + action + " " + details + "..."
    else:
        output = get_time_now() + " " + action + "..."

    print(output, end="")
    write_to_log(output)

def console_complete(result, status): # Status is a boolean representing success
    # E.g. Done.
    # Newline, green or red.

    output = " " + result + "."
    print(output)
    write_to_log(output + "\n")

def console_message(message):
    
    # Newline, magenta.
    output = get_time_now() + " " + message + "."
    print(output)
    write_to_log(output + "\n")


def find_match(sheet, column, text): # Search the existing data for matches... if found, don't count this as a new listing # Sheet must be an Excel worksheet, Column should be in set {A,B,C,D,E} and Text should be alphanumeric. # Returns True if a match is found.

    if text is not None: # Check that the string provided exists

        # Get all cells from specified column
        for cell in sheet[column]:

            if(cell.value is not None): # We need to check that the cell is not empty

                if cell.value == text: # Check if the value of the cell contains the text

                    return True # return true then break out of this function
        
    return False # Else (if none found) return false


# Define variables

# File version information
version = "1.8.1"
year_copyright = "2020"

# Email details
email_sender = 'newscrape.listings@gmail.com'
email_password = 'lettuce5' # If the credentials are denied, go to this address to: https://myaccount.google.com/lesssecureapps?pli=1
email_self = 'newscrape.listings@gmail.com'
email_client = "julescarboni013@gmail.com" #'cam@camilloventura.com'

# Root password for sudo
root_password = "lettuce5"

# VPN data
vpn_country_code = "au"
vpn_server_offset = 210 # Added to server index to bring into working range (au204 to au512)
vpn_server_maximum = 300 # Maximum number of servers that exist for this country (must be offset for use).
# rest of vpn variables declared below

# File paths
logs_path = r"/home/webscraper/Documents/Newscrape/Logs/"
results_path = r"/home/webscraper/Documents/Newscrape/Results/"

# Data to use to get listings
#clues = list_data.get_clues() # Clues (as a tuple)
clues = [] # Clues is an empty list so the user can specifiy which clue to get (a feature in testing)
states = list_data.get_states() # States (as a tuple)
suburbs = list_data.get_suburbs() # Suburbs (as a dictionary of tuples)

# Time calculation estimates
time_per_search = 9.838699106057437 # Experimental value
time_per_check = 0.8423052451580105 # Experimental value
time_per_rank = 0.0 #TEMPVAR (ASIC RANKING) # CURRENTLY 0 BECAUSE NOT IMPLEMENTED

pages_multiplier = 1.027971 # Add 2.8% (derived from test data)
checks_multiplier = 4.273
ranks_multiplier = 0.05

# Web address template
address_base = ("https://www.yellowpages.com.au/search/listings?clue=", "&eventType=pagination&openNow=false&pageNumber=", "&referredBy=UNKNOWN&&state=", "&suburb=", "+") # [0], clue, [1], page, [2] state, [3] suburb+state)
# Set capabilities of the firefox driver (specifically that it is allowed to ignore SSL/insecurity warnings)
browser_capabilities = DesiredCapabilities.FIREFOX.copy()
browser_capabilities['accept_untrusted_certs'] = True



# BEGIN ACTIONS


print("\n" + "NEWSCRAPE – Yellow Pages Web-Scraper", \
    "Jules Carboni, Copyright " + year_copyright + ", Version " + version, \
    sep="\n")

input("\n" + "Press Enter/Return key to begin (wait for password prompt please)...")



# Get current time for uptime calculations and count runs completed

time_start_uptime = datetime.now()

runs_completed = -1
runs_successful = 0

vpn_server_index = random.randint(0, vpn_server_maximum) # Randomly get for the current vpn server in use (can change mid-run)
vpn_forced_changes = 0



# TEMPORARY? Get clue to search.

clue_initial = input("Select a clue to use (e/p/c): ").lower()

if clue_initial == "e":
    this_clue = "electricians+electrical+contractors"
elif clue_initial == "p":
    this_clue = "plumbers+gas+fitters"
elif clue_initial == "c":
    this_clue = "builders+building+contractors"
else:
    console_message("Invalid clue, quitting Newscrape")
    exit()

clues.append(this_clue)




# This loop runs forever (until interrupted via keyboard)
while True:
    
    
    runs_completed += 1

    if 0 <= vpn_server_index < vpn_server_maximum: # If the index is between 0 to Max (300), then increment it to the next index, else reset it to zero.
        vpn_server_index += 1
    else:
        vpn_server_index = 0


    # Get timedelta for uptime
    time_start_run = datetime.now()
    time_uptime = time_start_run - time_start_uptime # Final minus initial
    

    print()
    print("Running Newcrape...")
    print("Run: " + str(runs_completed + 1) + " (" + str(runs_successful) + " successful)")
    print("Uptime: " + str(time_uptime) + " (" + str(time_uptime.days) + " days)")
    print()
    
    
    # Try to complete a run, if failed (expt. keyb-int): notify self by email (incl. error message) and start the next run.
    try:
        


        # Declare session specific variables


        # Get current time
        time_atm = datetime.now() # Get time stamp for output files
        time_atm_string = time_atm.strftime("%d%m%Y_%H%M%S")

        # Get current session's file names
        all_file = "all_results.xlsx"
        all_backup_file = "all_results_" + time_atm_string + ".xlsx.bak"
        new_file = "new_results_" + time_atm_string + ".xlsx"
        template_file = "new_results_template.xlsx"
        # Get current session's file paths
        all_path = results_path + all_file
        all_backup_path = results_path + all_backup_file
        new_path = results_path + new_file
        template_path = results_path + template_file


        # Declare list(s) that will hold listing records.
        listings = []
        #new_listings = []



        # Make a new log file for this session

        if not os.path.exists(logs_path): # Create the directories if they don't exist
            os.makedirs(logs_path)

        # Make filename and path
        log_file_name = "newscrape_" + time_atm_string + ".log"
        log_file_path = os.path.join(logs_path, log_file_name)

        with open(log_file_path, 'w') as log_file:
            log_file.write('Newscrape Console Log ' + get_time_now() + " (Jules Carboni, Copyright 2020, Version " + version + ")\n") # Write the header

        console_message("Created new log file " + log_file_name)



        # Count total pages to search through

        total_clues = len(clues)

        total_suburbs = 0
        for state in states:
            total_suburbs += len(suburbs[state]) # Add up all the suburbs
        total_searches = total_suburbs * total_clues # Multiply by the number of times we have to fetch from each suburb
        expected_searches = total_searches * pages_multiplier # Multiply to get add the amount of pages we expect to be in total
        checks_per_clue = checks_multiplier * total_suburbs * pages_multiplier # A really rough average
        ranks_per_clue = ranks_multiplier * total_suburbs * pages_multiplier #TEMPVAR (ASIC RANKING)
        total_checks = checks_per_clue * total_clues
        total_ranks = ranks_per_clue * total_clues

        expected_duration = expected_searches * time_per_search + total_checks * time_per_check + total_ranks * time_per_rank # Calculate the expected total duration of the program
        expected_duration_timedelta = timedelta(seconds=expected_duration)

        start_time = datetime.now()

        # Print this information and time estimates

        console_message("Calculated expected results..." + "\n" \
            "Total suburbs to search: " + str(total_suburbs) + "\n" \
            "Total searches to do: " + str(total_searches) + "\n" \
            "Expected total pages to search: " + str(int(expected_searches)) + "\n" \
            "Expected run duration: " + format_time_seconds(expected_duration) + " (" + str(expected_duration_timedelta.days) + " days)" + "\n" \
            "Expected time of completion: " + str(start_time + expected_duration_timedelta) \
            )



        # Make a fresh copy of the new-results template for editing.
        console_action("Making new spreadsheets", "")
        copyfile(template_path, new_path) # Create new blank sheet for new listings
        copyfile(all_path, all_backup_path) # Create a backup of the current sheet of all listings
        console_complete("Done", True)


        # Close any existing VPN services.
        console_action("Quitting existing VPN connections", "")
        #os.system("sudo openpyn -x")
        this_command = "sudo openpyn -x"
        call('echo {} | sudo -S {}'.format(root_password, this_command), shell=True)
        console_complete("Done", True)

        # Start the VPN service and log into a server.
        this_vpn_server = vpn_country_code + str(vpn_server_index + vpn_server_offset)
        console_action("Connecting to VPN server", this_vpn_server.upper())
        #os.system("sudo openpyn -s " + this_vpn_server + " --daemon")
        this_command = "sudo openpyn -s " + this_vpn_server + " --daemon"
        call('echo {} | sudo -S {}'.format(root_password, this_command), shell=True)
        console_complete("Done", True)


        # Initialise the selenium webdriver
        console_action("Initialising Selenium webdriver", "")
        browser = webdriver.Firefox(capabilities=browser_capabilities)
        console_complete("Done", True)



        # Download and extract data from every page!

        # Initialise counting variables
        suburbs_counter = 0
        pages_counter = 0

        # Loop through each, clue, state, suburb, and page.

        try:

            for clue in clues:

                for state in states:

                    for suburb in suburbs[state]: # Only the suburbs in this state

                        suburbs_counter += 1

                        # Reset page numbers file
                        pages_number = -1

                        # Loop through all pages
                        page = 1
                        while ( (page <= pages_number or pages_number == -1) and page <= 8):

                            print() # Add an empty line (to the terminal printout only) to make it more readable at a glance # DEBUG

                            pages_counter += 1


                            # Construct the url
                            web_address = address_base[0] + clue + address_base[1] + str(page) + address_base[2] + state + address_base[3] + suburb + address_base[4] + state


                            # Get the html for this page
                            while True: # Loop forever until successful

                                try:

                                    # Get time values for the console display
                                    percentage_complete = (suburbs_counter / total_searches) * 100
                                    time_remaining = expected_duration - (suburbs_counter * time_per_search)
                                    #time_remaining_timedelta = timedelta(seconds=time_remaining)

                                    progress_stamp = "[" + str(round(percentage_complete, 2)) + "% T-" + format_time_seconds(time_remaining) + "] "
                                    console_action(progress_stamp + "Getting", web_address)

                                    browser.get(web_address) # Get the webpage from the Internet
                                    #WebDriverWait(driver, 30).until(readystate_complete) # Wait until the page is fully loaded

                                    html_source = browser.page_source # Get the source from the browser driver

                                    console_complete("Done", True)



                                    # Extract the HTML

                                    console_action("Parsing and extracting HTML code", "")

                                    soup = BeautifulSoup(html_source, 'html.parser')
                                    listings_html = soup.find_all("div", attrs={"class": "listing listing-search listing-data"})

                                    console_complete("Done", True)



                                    # Check that the page isn't a "no robots" block page.

                                    console_action("Checking if current IP is blocked", "")

                                    # Try to get the html data (if can't get it, then its all fine so give up).
                                    try:
                                        robot_check_parent = soup.find("div", attrs={"style": "padding-top: 10px;"})
                                        robot_check_html = robot_check_parent.find("h1", attrs={"style": "font-weight: normal;"})
                                        robot_check_text = robot_check_html.text

                                        if "detected unusual traffic" in robot_check_text.lower():

                                            console_complete("Blocked", False)

                                            # Restart the vpn service with a new server.

                                            # Get new index (THIS IS REPEATED CODE! CONSOLIDATE INTO A FUNCTION LATER)     
                                            if 0 <= vpn_server_index < vpn_server_maximum:
                                                vpn_server_index += 1
                                            else:
                                                vpn_server_index = 0

                                            # Connect to the new server
                                            this_vpn_server = vpn_country_code + str(vpn_server_index + vpn_server_offset)

                                            console_action("Connecting to new VPN server", this_vpn_server.upper())

                                            #os.system("sudo openpyn -s " + this_vpn_server + " --daemon")
                                            this_command = "sudo openpyn -s " + this_vpn_server + " --daemon"
                                            call('echo {} | sudo -S {}'.format(root_password, this_command), shell=True)

                                            console_complete("Done", True)

                                            # Update forced vpn change counter.
                                            vpn_forced_changes += 1

                                        else: # If did not find "blocked message" (but also didnt encounter any errors), break out of the loop

                                            console_complete("Connected", False)

                                            break

                                    except:

                                        # If all fine (not blocked), don't try to get the page again (break out).

                                        console_complete("Connected", True)

                                        break


                                #except KeyboardInterrupt:

                                #    console_complete("Skipped", False)

                                #    # Stop trying to download this page and skip the rest of this iteration (return to top of outer while loop)
                                #    trying_saving = False

                                except Exception as ex:

                                    console_complete("Failed (" + str(ex) + ")", False) # Display error message (details)


                                    # Reinitialise the selenium webdriver to circumvent certificate errors
                                    console_action("Restarting web driver", "")

                                    browser.quit() # Gracefully quit driver
                                    browser = webdriver.Firefox(capabilities=browser_capabilities) # Restart the driver

                                    console_complete("Done", True)

                                    # Wait ten seconds before retrying
                                    #sleep(10)



                            

                            # If the number of pages hasn't been retrieved, get it from the page just downloaded
                            if pages_number == -1:

                                console_action("Finding number of pages", "")

                                try:

                                    # Extract the number of results
                                    results_number_parent = soup.find("div", attrs={"class": "cell search-message first-cell"})
                                    results_number_html = results_number_parent.find("span", attrs={"class": "emphasise"})
                                    results_number = int(results_number_html.text.split()[0])

                                    # Do the math
                                    if results_number >= 1 and results_number <= 1500:

                                        pages_number = int(ceil(results_number / 35.0))

                                        console_complete("Done", True)

                                    else:

                                        pages_number = -1

                                        console_complete("Unreasonable result", False)

                                except:

                                    pages_number = 0

                                    console_complete("Failed (assuming zero results)", False)


                            # Gather all listings' data and check if they have already been identified.

                            for listing_html in listings_html:

                                console_action("Extracting listing", listing_html.get('data-listing-name'))

                                ## Make links absolute
                                base_url = "https://www.yellowpages.com.au"
                                #listing_html.make_links_absolute(base_url)


                                # Reset variables
                                business_name_html = None
                                phone_number_html = None
                                email_address_html = None
                                business_website_html = None
                                name = None
                                phone = None
                                email = None
                                website = None
                                yellow_page = None

                                # Get the html data of each match.
                                business_name_html = listing_html.find("a", attrs={"class": "listing-name"})
                                #phone_number_html = listing_html.find("a", attrs={"class": "click-to-call contact contact-preferred contact-phone"})
                                phone_number_html = listing_html.find("a", attrs={"class": lambda x: x and 'click-to-call contact contact-preferred' in x})
                                email_address_html = listing_html.find("a", attrs={"class": "contact contact-main contact-email"})
                                business_website_html = listing_html.find("a", attrs={"class": "contact contact-main contact-url"})


                                # Extract the appropriate sections of each html element.

                                # Get sections (only if they exist)
                                if business_name_html is not None:
                                    name = business_name_html.text
                                if phone_number_html is not None:
                                    phone = phone_number_html.get('href')
                                if email_address_html is not None:
                                    email = email_address_html.get('data-email')
                                if business_website_html is not None:
                                    website = business_website_html.get('href')
                                if business_name_html is not None:
                                    yellow_page = business_name_html.get('href') # Link to the Yellow Pages listing.

                                # Add sections to a record
                                record = listing(name, phone, email, website, base_url + yellow_page, state)
                                #record = listing(name, phone, email, base_url + yellow_page) # Use the YP listing instead of the actual website temporarily.
                                # Append this record to the list of listings
                                listings.append(record)


                                console_complete("Done", True)


                            page += 1 # Increment page number

        except Exception as ex:

            # If any shit hits the fan, stop searching and move onto checking so that we at least don't lose the listings we have

            console_message("Searching interrupted unexpectedly (" + str(ex) + ")")


        # All done getting pages. Gracefully quit the Firefox browser and the Selenium web driver.
        browser.quit()
        #webdriver.Firefox.quit


        time_searching_done = datetime.now() # The time at which all the downloading and searching was completed



        # Check if each business listing already exists in our local database

        try:

            print() # Add an empty line (to the terminal printout only) to make it more readable at a glance # DEBUG


            # Load all-results worksheet

            console_action("Loading worksheets", "")

            book_all = load_workbook(filename = all_path) # Load the workbook
            sheet_all = book_all['Sheet1'] # Load the worksheet

            book_new = load_workbook(filename = new_path) # Load the workbook
            sheet_new = book_new['Sheet1'] # Load the worksheet

            console_complete("Done", True)


            # Get the last row in the sheet
            final_row_all = sheet_all.max_row
            final_row_new = 1 #sheet_new.max_row # Always start at the top of the sheet


            # Calculate the expected time remaining
            expected_duration_remaining = expected_duration - (expected_searches * time_per_search)


            # Loop through each newly retrieved record

            index = 0
            # Use seperate indexes for sheets so that no rows are skipped in the spreadsheet
            sheet_index_all = final_row_all
            sheet_index_new = final_row_new

            listings_count = len(listings) # For each listing gathered
            while index <= listings_count - 1:

                business = listings[index]

                # Get time values for the console display
                decimal_complete = index / listings_count
                percentage_complete = (decimal_complete) * 100
                #time_remaining = expected_duration_remaining - (index * time_per_check)
                time_remaining = expected_duration_remaining * (1 - decimal_complete)

                progress_stamp = "[" + str(round(percentage_complete, 2)) + "% T-" + format_time_seconds(time_remaining) + "] "
                console_action(progress_stamp + "Checking", business.business_name)

                this_name = business.business_name
                this_phone = business.phone_number
                this_email = business.email_address
                this_website = business.business_website
                this_yellow_page = business.yellow_pages_link
                this_location_state = business.location_state

                # Search the sheet of all listings to see if it already exists
                #this_name_exists = find_match(sheet_all, "A", this_name)
                this_name_exists = False # ALWAYS RETURN FALSE, WE WANT TO IGNORE THIS CHECK
                this_phone_exists = find_match(sheet_all, "B", this_phone)
                this_email_exists = find_match(sheet_all, "C", this_email)
                this_website_exists = find_match(sheet_all, "D", this_website)
                this_yellow_page_exists = find_match(sheet_all, "E", this_yellow_page)

                if this_name_exists or this_phone_exists or this_email_exists or this_website_exists or this_yellow_page_exists: # If any of the searches returned a True result for existence

                    console_complete("Already exists", False)

                    # Keep sheet indexes the same

                else:

                    console_complete("Found new listing", True)

                    # Add the new listing to all-results spreadsheet

                    # Using the old max row as a base get the next row number to write to
                    # AKA Increment the indexes for this new row
                    sheet_index_all += 1
                    sheet_index_new += 1

                    console_action("Adding", business.business_name + " to spreadsheets")


                    # Add to all listings sheet
                    sheet_all.cell(row=sheet_index_all, column=1).value = this_name # Name
                    sheet_all.cell(row=sheet_index_all, column=2).value = this_phone # Phone
                    sheet_all.cell(row=sheet_index_all, column=3).value = this_email # Email
                    sheet_all.cell(row=sheet_index_all, column=4).value = this_website # Website
                    #sheet_all.cell(row=sheet_index_all, column=5).value = this_yellow_page # Yellow Page


                    ##console_action("Getting ASIC data for", business.business_name)

                    ####### GET ASIC DATA!!!!!!!!!!!!!!!!!!
                    # GET ASIC (ABN/ACN) DETAILS and SORT NEW LISTINGS BASED ON IF THEY HAVE A WEBSITE/ABN/ACN ((SEE ABOVE--GOES INSIDE FOR LOOP)).
                    # ADD ASIC RANKING TO TIME CALCULATIONS
                    #### Use the data in an algorithm to produce a ranking!!!


                    # Add to new listings sheet
                    sheet_new.cell(row=sheet_index_new, column=1).value = this_name # Name
                    sheet_new.cell(row=sheet_index_new, column=2).value = this_phone # Phone
                    sheet_new.cell(row=sheet_index_new, column=3).value = this_email # Email
                    sheet_new.cell(row=sheet_index_new, column=4).value = this_website # Website
                    #sheet_new.cell(row=sheet_index_new, column=5).value = this_yellow_page # Yellow Page
                    sheet_new.cell(row=sheet_index_new, column=5).value = this_location_state # Location State


                    # Save the changes to the files
                    book_all.save(all_path)
                    book_new.save(new_path)


                    console_complete("Done", True)


                index += 1 # Increment index (b/c not using a for loop)


        except Exception as ex:

            # If any shit hits the fan, stop checking and move onto emailing so that we at least don't lose the listings we have

            console_message("Checking interrupted unexpectedly (" + str(ex) + ")")


        # Save all changes
        console_action("Saving spreadsheets", "") # ALWAYS DO THIS

        # Save the changes to the files
        book_all.save(all_path)
        book_new.save(new_path)

        book_all.close()

        console_complete("Done", True)


        print() # Add an empty line (to the terminal printout only) to make it more readable at a glance # DEBUG


        time_checking_done = datetime.now() # The time at which all the checking was completed (incl. writing to the all-spreadsheet)

        #time_ranking_done = datetime.now() # The time at which all the ranking (with ASIC) was completed (incl. writing to the new spreadsheet)



        # Calculate statistics

        #total_pages_count = pages_counter
        total_listings_count = len(listings)
        #new_listings_count = len(new_listings)
        new_listings_count = sheet_index_new - final_row_new
        #all_listings_count = sheet_index_all - final_row_all #same as tot_list_count

        finish_time = datetime.now()
        total_duration = finish_time - start_time
        difference_duration = total_duration - expected_duration_timedelta

        time_searching_duration = time_searching_done - start_time
        time_checking_duration = time_checking_done - time_searching_done
        #time_ranking_duration = time_ranking_done - time_checking_done

        time_per_search_actual = time_searching_duration.total_seconds() / expected_searches
        time_per_check_actual = time_checking_duration.total_seconds() / total_listings_count
        #time_per_rank_actual = time_ranking_duration.total_seconds() / new_listings_count



        # Update the calculation times by averaging the expected and actual values
        time_per_search = (time_per_search + time_per_search_actual) / 2
        time_per_check = (time_per_check + time_per_check_actual) / 2
        #time_per_rank = (time_per_rank + time_per_rank_actual) / 2


        # Prediction for duration of next run
        next_expected_duration = expected_searches * time_per_search + total_checks * time_per_check + total_ranks * time_per_rank # Calculate the expected total duration of the program
        next_expected_duration_timedelta = timedelta(seconds=next_expected_duration)
        next_finish_time = finish_time + next_expected_duration_timedelta
        next_finish_time_string = next_finish_time.strftime("%c")
        #next_finish_time_string = str(next_finish_time)


        # Get timedelta for uptime
        time_uptime = finish_time - time_start_uptime # Final minus initial
    

        status_message = ("Newscrape – Results Report" + "\n"
                            "\n"
                            "Version: " + version + "\n"
                            "Time: " + finish_time.strftime("%c") + "\n"
                            "Uptime: " + str(time_uptime) + " (" + str(time_uptime.days) + " days)" + "\n"
                            "Run: " + str(runs_completed + 1) + " (" + str(runs_successful) + " successful)" + "\n"
                            "VPN: " + vpn_country_code + str(vpn_server_index + vpn_server_offset) + " (" + str(vpn_forced_changes) + " forced changes)" + "\n")

        results_message = ("Total pages: " + str(pages_counter) + "\n" \
            "Total listings: " + str(total_listings_count) + "\n" \
            "Total new listings: " + str(new_listings_count) + "\n" \
            "Duration of searching: " + str(time_searching_duration) + "\n" \
            "Duration of checking: " + str(time_checking_duration) + "\n" \
            #"Duration of ranking: " + str(time_ranking_duration) + "\n" \
            "Total duration: " + str(total_duration) + "\n" \
            "Expected duration: " + str(expected_duration_timedelta) + "\n" \
            #"Difference from expected: " + difference_duration.strftime("%H:%M:%S") + "\n" \
            "Difference from expected: " + str(difference_duration) + "\n" \
            "Time per search: " + str(time_per_search_actual) + "\n" \
            "Time per check: " + str(time_per_check_actual) \
            #"Time per rank: " + str(time_per_rank_actual) + "\n" \
        )

        next_finish_message = "New list expected by: " + next_finish_time_string
        


        # EMAIL THE SHEETS.


        for email_recipient in (email_self, email_client):

            console_action("Emailing results", "to " + email_recipient)

            try:

                # instance of MIMEMultipart 
                msg = MIMEMultipart() 

                # storing the senders email address   
                msg['From'] = email_sender 
                # storing the receivers email address  
                msg['To'] = email_recipient 



                # string to store the body of the mail

                if email_recipient == email_self:
                    body = status_message + "\n" + results_message + "\n" + next_finish_message
                else:
                    body = "Number of new listings: " + str(new_listings_count) + "\n\n" + next_finish_message

                # attach the body with the msg instance 
                msg.attach(MIMEText(body, 'plain')) 


                # storing the subject  
                start_time_string = start_time.strftime("%d %b")
                emailing_time_string = datetime.now().strftime("%d %b")
                msg['Subject'] = 'New Listings from ' + start_time_string + ' to ' + emailing_time_string



                # open the file to be sent  
                filename = new_file
                attachment = open(new_path, "rb") 


                # instance of MIMEBase and named as mime
                mime = MIMEBase('application', 'octet-stream') 
                # To change the payload into encoded form 
                mime.set_payload((attachment).read()) 
                # encode into base64 
                encoders.encode_base64(mime) 
                mime.add_header('Content-Disposition', "attachment; filename= %s" % filename) 
                # attach the instance 'mime' to instance 'msg' 
                msg.attach(mime)
                # creates SMTP session 
                smtp = smtplib.SMTP('smtp.gmail.com', 587) 
                # start TLS for security 
                smtp.starttls() 

                # Authentication 
                smtp.login(email_sender, email_password) 

                # Converts the Multipart msg into a string 
                text = msg.as_string() 

                # sending the mail 
                smtp.sendmail(email_sender, email_recipient, text) 


                console_complete("Success", True)

            except Exception as ex:

                console_complete("Failed (" + str(ex) + ")", False)


            # terminating the session (ALWAYS DO THIS)
            smtp.quit()



        # Update number of completed runs
        runs_successful += 1

        print() # Add an empty line (to the terminal printout only) to make it more readable at a glance # DEBUG

        console_message("Updated successful runs counter (" + str(runs_successful) + " successful)")



        # Print stats

        console_message("Calculated actual results..." + "\n" + results_message)
        
        
        
    # If the run fails!:
    
    except KeyboardInterrupt:
            
        # Ask if the interruptor would like to quit the program immediately OR start the next run.
        
        print("\n")

        # Try to write this action to file (if the file exists basically).
        try:
            console_message("Keyboard interrupt detected")
        except:
            print("Keyboard interrupt detected.")

        choice = input("To quit Newscrape enter \"quit\", or enter anything else to begin the next run: ")
        
        if choice.lower() in ("q", "quit") : # Only exit if above condition met.
            try:
                console_message("Quit Newscrape manually")
            except:
                pass

            exit() # Quit (close selenium/browser but not openvpn/nordvpn)


    except Exception as run_ex:
        
        #console_message("") # Always print the error message on a new line.
        print("\n") # Add an empty line (to the terminal printout only) to make it more readable at a glance # DEBUG

        # For any other exception, email the error code to self, then start the next run.
        
        crash_time = datetime.now() # Get the time of the crash
        crash_time_string_short = crash_time.strftime("%d/%m/%Y %H:%M:%S")
        crash_time_string_long = crash_time.strftime("%c")

        try:
            console_message("Caught error (" + str(run_ex) + ")")
        except:
            print("Caught error (" + str(run_ex) + ").")


        try:
            console_action("Emailing crash report", "to " + email_self)
        except:
            pass
    
        try: # Try to send the email, if it doesn't work, just go to the next run

            # instance of MIMEMultipart 
            msg = MIMEMultipart() 

            # storing the senders email address   
            msg['From'] = email_sender 
            # storing the receivers email address  
            msg['To'] = email_self 


            
            # string to store the body of the mail
            
            
            # Get timedelta for uptime
            time_crash_run = datetime.now()
            time_uptime = time_crash_run - time_start_uptime # Final minus initial


            # Get error message details
            error_message = repr(run_ex)
        
            body_message = ("Newscrape – Crash/Error Report" + "\n"
                            "\n"
                            "Version: " + version + "\n"
                            "Time: " + crash_time_string_long + "\n"
                            "Uptime: " + str(time_uptime) + " (" + str(time_uptime.days) + " days)" + "\n"
                            "Run: " + str(runs_completed + 1) + " (" + str(runs_successful) + " successful)" + "\n"
                            "VPN: " + vpn_country_code + str(vpn_server_index + vpn_server_offset) + " (" + str(vpn_forced_changes) + " forced changes)" + "\n"
                            "\n"
                            "Exception details:" + "\n"
                            "" + error_message + "\n")
    
            
            body = body_message
            
            # attach the body with the msg instance 
            msg.attach(MIMEText(body, 'plain')) 

            
            
            # storing the subject
            msg['Subject'] = 'Crash Report ' + crash_time_string_short


            
            # open the file to be sent  
            filename = log_file_name
            filepath_temp = logs_path + log_file_name
            attachment = open(filepath_temp, "rb") 

            # instance of MIMEBase and named as mime
            mime = MIMEBase('application', 'octet-stream') 
            # To change the payload into encoded form 
            mime.set_payload((attachment).read()) 
            # encode into base64 
            encoders.encode_base64(mime) 
            mime.add_header('Content-Disposition', "attachment; filename= %s" % filename) 
            # attach the instance 'mime' to instance 'msg' 
            msg.attach(mime)
            # creates SMTP session 
            smtp = smtplib.SMTP('smtp.gmail.com', 587) 
            # start TLS for security 
            smtp.starttls() 

            # Authentication 
            smtp.login(email_sender, email_password) 

            # Converts the Multipart msg into a string 
            text = msg.as_string() 

            # sending the mail 
            smtp.sendmail(email_sender, email_self, text) 


            try:
                console_complete("Success", True)
            except:
                print("Successfully sent error report.")

        except Exception as ex:

            try:
                console_complete("Failed (" + str(ex) + ").", False)
            except:
                print("Failed to send error report (" + str(ex) + ").")



        # terminating the session (ALWAYS DO THIS)
        smtp.quit()
